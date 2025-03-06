import os
import random
import string
import tempfile
from datetime import datetime, timedelta
import io

import pytest
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from playwright.sync_api import sync_playwright, expect, Page
from time import sleep


# Define waste materials by category
WASTE_MATERIALS = {
    "General Waste": ["Mixed Waste", "Office Waste", "Food Packaging", "Non-recyclable Plastics"],
    "Recyclable": ["Cardboard", "Paper", "Aluminum Cans", "Plastic Bottles", "Glass Bottles"],
    "Hazardous": ["Electronic Waste", "Batteries", "Chemical Containers", "Paint Cans"],
    "Organic": ["Food Waste", "Garden Waste", "Wood Chips", "Compostable Materials"]
}

# Category mapping for web application
CATEGORY_MAPPING = {
    "General Waste": "1",
    "Recyclable": "2",
    "Hazardous": "3",
    "Organic": "4"
}


# Utility functions
def random_string(length=8):
    """Generate a random string for invoice IDs."""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))


def random_weight():
    """Generate a random weight in kg (2 decimal places)."""
    return round(random.uniform(50.0, 10000.0), 2)


def random_unit_price():
    """Generate a random unit price (2 decimal places)."""
    return round(random.uniform(0.5, 10.0), 2)


def calculate_total(weight, unit_price):
    """Calculate total price based on weight and unit price."""
    return round(weight * unit_price, 2)


def get_random_date(days_back=30):
    """Generate a random date within the last 30 days."""
    today = datetime.today()
    random_days = random.randint(1, days_back)
    random_date = today - timedelta(days=random_days)
    return random_date.strftime("%Y-%m-%d")


def prepare_excel_template(template_path, output_path, data):
    """Prepare the Excel template with the provided data."""
    # Load the workbook
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # Fill in the data (adjust cell references based on actual template)
    sheet['B4'] = data['company_name']  # Vendor name
    sheet['B6'] = data['client_name']   # Client name
    
    # Description row - material and weight
    description_row = 12  # Adjust this to match your template
    description = f"{data['material']} - {data['weight']} kg"
    sheet.cell(row=description_row, column=2).value = description
    
    # Unit price, quantity, and total
    sheet.cell(row=description_row, column=3).value = data['weight']
    sheet.cell(row=description_row, column=4).value = data['unit_price']
    sheet.cell(row=description_row, column=5).value = data['total']
    
    # Invoice number
    sheet['E4'] = data['invoice_id']  # Invoice ID
    
    # Save the modified workbook
    wb.save(output_path)
    return output_path


def convert_excel_to_pdf(excel_path, pdf_dir, pdf_name, invoice_data):
    """Create a simple but valid PDF invoice using ReportLab canvas directly."""
    pdf_path = os.path.join(pdf_dir, pdf_name)
    
    try:
        # Create a simple PDF using canvas directly (more reliable)
        c = canvas.Canvas(pdf_path, pagesize=letter)
        
        # Add invoice information
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 750, f"INVOICE #{invoice_data['invoice_id']}")
        
        c.setFont("Helvetica", 12)
        c.drawString(100, 720, f"Vendor: {invoice_data['company_name']}")
        c.drawString(100, 700, f"Client: {invoice_data['client_name']}")
        c.drawString(100, 680, f"Date: {datetime.now().strftime('%Y-%m-%d')}")
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(100, 650, "Description")
        c.drawString(350, 650, "Weight (kg)")
        c.drawString(430, 650, "Unit Price")
        c.drawString(500, 650, "Amount")
        
        # Draw a line
        c.line(100, 640, 500, 640)
        
        # Add invoice item
        c.setFont("Helvetica", 12)
        c.drawString(100, 620, f"{invoice_data['material']}")
        c.drawString(350, 620, f"{invoice_data['weight']:.2f}")
        c.drawString(430, 620, f"${invoice_data['unit_price']:.2f}")
        c.drawString(500, 620, f"${invoice_data['total']:.2f}")
        
        # Draw another line
        c.line(100, 610, 500, 610)
        
        # Add total
        c.setFont("Helvetica-Bold", 12)
        c.drawString(430, 590, "Total:")
        c.drawString(500, 590, f"${invoice_data['total']:.2f}")
        
        # Add waste category info
        c.setFont("Helvetica", 12)
        c.drawString(100, 550, f"Waste Category: {invoice_data['category']}")
        
        # Finalize the PDF
        c.save()
        
        print(f"PDF created successfully at {pdf_path}")
        
    except Exception as e:
        print(f"PDF creation failed: {e}")
        print("Creating a minimal PDF file for testing...")
        
        # Create a minimal PDF file for testing that is guaranteed to work
        c = canvas.Canvas(pdf_path)
        c.drawString(100, 750, "Test Invoice")
        c.save()
    
    return pdf_path


def generate_invoice_data():
    """Generate random invoice data."""
    # First select a category, then a material from that category
    category = random.choice(list(WASTE_MATERIALS.keys()))
    material = random.choice(WASTE_MATERIALS[category])
    
    weight = random_weight()
    unit_price = random_unit_price()
    total = calculate_total(weight, unit_price)
    invoice_id = f"INV-{random_string(6)}"
    
    return {
        'company_name': "Demo Vendor",
        'client_name': "ABC Company Pte Ltd",
        'category': category,
        'material': material,
        'weight': weight,
        'unit_price': unit_price,
        'total': total,
        'invoice_id': invoice_id
    }


# Fixtures
@pytest.fixture(scope="function")
def browser_context():
    """Fixture to set up and tear down the browser context."""
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context(viewport={"width": 2120, "height": 1360})
        page = context.new_page()
        try:
            yield page
        finally:
            context.close()
            browser.close()


@pytest.fixture(scope="function")
def excel_template():
    """Fixture to provide the Excel template path."""
    template_path = os.path.join(os.path.dirname(__file__), "template.xlsx")
    
    if not os.path.exists(template_path):
        pytest.skip(f"Excel template not found at {template_path}")
    
    return template_path


@pytest.fixture(scope="function")
def generated_pdf(request):
    """Fixture to generate a PDF with random invoice data."""
    # Generate random invoice data
    invoice_data = generate_invoice_data()
    
    # Create temporary directory for PDF
    temp_dir = tempfile.mkdtemp()
    
    # Create PDF filename
    pdf_filename = f"{invoice_data['invoice_id']}_{invoice_data['company_name']}_{invoice_data['material']}.pdf"
    pdf_filename = pdf_filename.replace(' ', '_')
    
    # Convert directly to PDF (skip Excel modification)
    pdf_path = convert_excel_to_pdf(None, temp_dir, pdf_filename, invoice_data)
    
    # Print for debugging
    print(f"Created PDF at: {pdf_path}")
    print(f"File exists: {os.path.exists(pdf_path)}")
    print(f"File size: {os.path.getsize(pdf_path)} bytes")
    
    # Yield both PDF path and invoice data
    yield {"path": pdf_path, "data": invoice_data}
    
    # Cleanup after test
    def teardown():
        import shutil
        shutil.rmtree(temp_dir)
    request.addfinalizer(teardown)


# Test functions
def test_upload_invoice_pdf(browser_context, generated_pdf):
    """Test uploading a generated PDF to the invoice management system."""
    page = browser_context
    pdf_path = generated_pdf["path"]
    invoice_data = generated_pdf["data"]
    invoice_id = invoice_data["invoice_id"]
    category = invoice_data["category"]
    
    # Get category ID for selection in UI
    category_id = CATEGORY_MAPPING.get(category, "1")
    
    try:
        # Navigate to the application
        page.goto("http://localhost:3000/")

        # Ensure page is fully loaded
        page.wait_for_load_state("networkidle")
        
        # Ensure we're on the upload tab
        page.get_by_role("button", name="Upload Files").click()
        
        # Fill invoice details
        page.get_by_role("textbox", name="Enter invoice number").click()
        page.get_by_role("textbox", name="Enter invoice number").fill(invoice_id)
        
        # Set invoice date
        random_date = get_random_date(30)
        page.locator("input[type=\"date\"]").fill(random_date)
        
        # Select waste category
        page.get_by_role("combobox").select_option(category_id)
        
        # Upload PDF file
        with page.expect_file_chooser() as file_chooser_info:
            page.get_by_text("Browse").click()
        file_chooser = file_chooser_info.value
        file_chooser.set_files(pdf_path)
        
        # Wait for file preview to appear
        preview = page.locator(".preview-section")
        expect(preview).to_be_visible(timeout=10000)
        
        # Submit the form
        upload_button = page.get_by_role("button", name="Upload Invoice")
        expect(upload_button).to_be_enabled()

        page.wait_for_load_state("networkidle")
        upload_button.click()
        
        # Wait for success notification
        page.wait_for_selector("text=Success: Invoice uploaded", timeout=20000)
        
        # Verify upload by checking status tab
        page.get_by_role("button", name="Status Review").click()
        
        # Use a more specific selector that targets only the invoice number column
        # This fixes the strict mode violation
        expect(page.locator(f"tr:has-text('{invoice_id}') >> nth=0")).to_be_visible(timeout=25000)
        
        print(f"✅ Successfully uploaded and verified invoice: {invoice_id}")
        
    except Exception as e:
        # Take screenshot on failure
        page.screenshot(path=f"test_failure_{invoice_id}.png")
        raise e


@pytest.mark.parametrize("test_run", range(5))
def test_multiple_invoice_uploads(browser_context, excel_template, test_run):
    """Test uploading multiple different invoices."""
    page = browser_context
    
    # Generate random invoice data for this run
    invoice_data = generate_invoice_data()
    category_id = CATEGORY_MAPPING.get(invoice_data["category"], "1")
    
    try:
        # Create temporary files and directories
        with tempfile.TemporaryDirectory() as pdf_dir:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_excel:
                modified_excel_path = tmp_excel.name
            
            try:
                # Prepare the Excel file
                prepare_excel_template(excel_template, modified_excel_path, invoice_data)
                
                # Create PDF filename and convert
                pdf_filename = f"{invoice_data['invoice_id']}_{invoice_data['company_name']}_{invoice_data['material']}.pdf"
                pdf_filename = pdf_filename.replace(' ', '_')
                pdf_path = convert_excel_to_pdf(modified_excel_path, pdf_dir, pdf_filename, invoice_data)
                
                # Upload and test
                page.goto("http://localhost:3000/")
                page.wait_for_load_state("networkidle")
                page.get_by_role("button", name="Upload Files").click()
                page.get_by_role("textbox", name="Enter invoice number").fill(invoice_data["invoice_id"])
                page.locator("input[type=\"date\"]").fill(get_random_date(30))
                page.get_by_role("combobox").select_option(category_id)
                
                with page.expect_file_chooser() as file_chooser_info:
                    page.get_by_text("Browse").click()
                file_chooser = file_chooser_info.value
                file_chooser.set_files(pdf_path)
                
                sleep(3)
                page.wait_for_selector(".preview-section", timeout=10000)
                sleep(5)
                page.get_by_role("button", name="Upload Invoice").click()
                page.wait_for_selector("text=Success: Invoice uploaded", timeout=15000)

                sleep(3)
                
                # Verify in status tab
                page.get_by_role("button", name="Status Review").click()
                
                # Use a more specific selector that targets the row containing the invoice ID
                expect(page.locator(f"tr:has-text('{invoice_data['invoice_id']}') >> nth=0")).to_be_visible(timeout=20000)

                sleep(3)
                print(f"✅ Successfully uploaded invoice: {invoice_data['invoice_id']} (Run {test_run+1}/5)")
                
            finally:
                # Clean up
                if os.path.exists(modified_excel_path):
                    os.unlink(modified_excel_path)
                    
    except Exception as e:
        page.screenshot(path=f"test_failure_multiple_{test_run}.png")
        raise e


if __name__ == "__main__":
    pytest.main(["-xvs"])